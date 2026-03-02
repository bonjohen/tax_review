"""Parse IRS SOI Sales of Capital Assets (SOCA) Table 4.

SOCA Table 4: Short-Term and Long-Term Capital Gains and Losses,
by Selected Asset Type and Length of Time Held.

File format: .xlsx (openpyxl), one sheet with sub-tables 4A-4E:
  4A: All asset types
  4B: Corporate stock
  4C: Bonds and other securities
  4D: Real estate
  4E: Other asset types

Each sub-table has short-term and long-term sections with holding
period duration bins. All money values in thousands of dollars.

Columns (openpyxl 1-indexed):
  col 1 = Duration label
  col 2 = Gain transactions: Number
  col 3 = Gain transactions: Sales price
  col 4 = Gain transactions: Basis
  col 5 = Gain transactions: Gain (thousands)
  col 6 = Loss transactions: Number
  col 7 = Loss transactions: Sales price
  col 8 = Loss transactions: Basis
  col 9 = Loss transactions: Loss (thousands)
"""

import logging
import re
import sqlite3
from pathlib import Path

import openpyxl

from .db import insert_rows
from .parse_table_1x import _clean_cell, _money

logger = logging.getLogger(__name__)

# Sub-table offsets: (asset_type, title_row) — found by scanning for "Table 4X"
_ASSET_TYPES = [
    "all_assets", "corporate_stock", "bonds_securities",
    "real_estate", "other_assets",
]


def _find_subtable_rows(ws) -> list[tuple[str, int]]:
    """Find the starting row of each sub-table (4A-4E).

    Returns list of (asset_type, title_row).
    """
    subtables = []
    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if val and str(val).strip().startswith("Table 4"):
            subtables.append(r)
    result = []
    for i, row in enumerate(subtables):
        if i < len(_ASSET_TYPES):
            result.append((_ASSET_TYPES[i], row))
    return result


def _parse_subtable(ws, title_row: int, asset_type: str, year: int) -> list[dict]:
    """Parse one sub-table (e.g. 4A) for both ST and LT duration rows.

    Scans from title_row downward for 'Short-term transactions' and
    'Long-term transactions' sections, stopping at footnotes.
    """
    rows = []
    section = None  # 'short_term' or 'long_term'
    # Scan from title_row + 5 (past headers) until footnotes
    start = title_row + 6
    for r in range(start, ws.max_row + 1):
        label_raw = ws.cell(r, 1).value
        if label_raw is None:
            continue
        label = str(label_raw).strip()
        if not label:
            continue

        # Stop at footnotes/notes
        if label.startswith("[") or label.startswith("NOTE") or label.startswith("Source"):
            break

        label_lower = label.lower()

        # Section markers
        if "short-term transactions" in label_lower:
            section = "short_term"
            continue
        if "long-term transactions" in label_lower:
            section = "long_term"
            continue
        if section is None:
            continue

        # Skip "Total" rows — we compute our own totals
        if label_lower == "total":
            continue

        # Parse duration row
        n_gain = _clean_cell(ws.cell(r, 2).value)
        gain = _money(ws.cell(r, 5).value)
        n_loss = _clean_cell(ws.cell(r, 6).value)
        loss = _money(ws.cell(r, 9).value)
        sales_price = _money(ws.cell(r, 3).value)
        cost_basis = _money(ws.cell(r, 4).value)

        # Clean up the duration label
        duration = label.strip()

        rows.append({
            "year": year,
            "asset_type": asset_type,
            "holding_period": section,
            "holding_duration": duration,
            "number_of_gain_transactions": n_gain,
            "gross_sales_price": (sales_price or 0) + (_money(ws.cell(r, 7).value) or 0),
            "cost_basis": (cost_basis or 0) + (_money(ws.cell(r, 8).value) or 0),
            "gain_amount": gain,
            "loss_amount": loss,
            "net_gain_loss": (gain or 0) - (loss or 0),
        })

    return rows


def parse_soca_t4(filepath: Path, year: int) -> list[dict]:
    """Parse SOCA Table 4 — Gains by holding duration.

    Returns list of dicts with all duration rows across asset types.
    """
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    ws = wb.active

    subtables = _find_subtable_rows(ws)
    all_rows = []
    for asset_type, title_row in subtables:
        rows = _parse_subtable(ws, title_row, asset_type, year)
        all_rows.extend(rows)
        logger.debug(f"  SOCA T4 {asset_type}: {len(rows)} duration rows")

    wb.close()
    logger.info(f"SOCA Table 4 TY{year}: parsed {len(all_rows)} rows from {filepath.name}")
    return all_rows


def load_soca_t4(conn: sqlite3.Connection, filepath: Path, year: int) -> int:
    """Parse SOCA Table 4 and insert into raw_soca_t4."""
    rows = parse_soca_t4(filepath, year)
    return insert_rows(conn, "raw_soca_t4", rows)


# --- Holding period classification helpers -----------------------------------

# Duration labels that map to the "1 to 5 year" bucket (long-term)
_LT_1_TO_5_PATTERNS = [
    re.compile(r"under 18 months", re.I),
    re.compile(r"18 months under 2 years", re.I),
    re.compile(r"2 years under 3 years", re.I),
    re.compile(r"3 years under 4 years", re.I),
    re.compile(r"4 years under 5 years", re.I),
]

# Duration labels that map to the "5+ years" bucket
_LT_5_PLUS_PATTERNS = [
    re.compile(r"5 years under 10 years", re.I),
    re.compile(r"10 years under 15 years", re.I),
    re.compile(r"15 years under 20 years", re.I),
    re.compile(r"20 years or more", re.I),
]


def classify_holding_duration(duration_label: str) -> str | None:
    """Classify a SOCA duration label into a reform-relevant bucket.

    Returns: 'lt_1_to_5yr', 'lt_5yr_plus', or None (for unclassifiable rows
    like 'Period not determinable').
    """
    for pattern in _LT_1_TO_5_PATTERNS:
        if pattern.search(duration_label):
            return "lt_1_to_5yr"
    for pattern in _LT_5_PLUS_PATTERNS:
        if pattern.search(duration_label):
            return "lt_5yr_plus"
    return None
